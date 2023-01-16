import datetime
class WorksheetTable:

    header = dict()
    data = list()

    def __init__(self, workbookName, worksheetName):
        import datetime
        from openpyxl import load_workbook
        wb = load_workbook(workbookName)
        ws = wb[worksheetName]

        headerRowNum = 2
        columnsToKeep = [x for x, val in enumerate(ws[str(headerRowNum)], 1)]
        columnsToKeep.pop()
        rowMin = headerRowNum
        rowMax = len(ws['A'])
        colMin = 1
        colMax = max(columnsToKeep)
        data = []
        dataHeader = {}
        
        tempHead = [row for row in ws.iter_rows(headerRowNum,headerRowNum,colMin,colMax,True)]
        tempHead = list(tempHead[0])
        for colCount, value in enumerate(tempHead):
            x = value
            x = x.replace('(d)','')
            x = x.replace('(*)','')
            dataHeader.update({x : colCount})
        self.header = dataHeader

        for rowCount, row in enumerate(ws.iter_rows(rowMin, rowMax, colMin, colMax, True), rowMin):

            tempList = []
            for value in row:
                if rowCount == rowMin:
                    continue
                else:
                    tempList.append(value)
            data.append(tempList)
        data.pop(0)

        self.data = data

    # Iterate through columns and return optimal row widths based on string length
    #   returns list of int

    def get_ColumnWidths(self):
        dateFormat = '%x'
        tableColWidths = []

        for row in self.data:
            for count, value in enumerate(row):
                x = value
                if type(x) == type(datetime.datetime.now()):
                    x = value.strftime(dateFormat)

                if len(str(x)) > tableColWidths[count]:
                    tableColWidths[count] = len(str(x))
        return tableColWidths

